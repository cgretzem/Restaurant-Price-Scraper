from openpyxl import load_workbook


class Excel:
    def __init__(self, filename):
        self.workbook = load_workbook(filename)

    def get_sheet_names(self):
        return self.workbook.sheetnames

    def get_all_zips(self):
        zips = []
        for sheet_name in self.workbook.sheetnames:
            self.set_active_sheet(sheet_name)
            ws_zips = self.read_zips()
            for code in ws_zips:
                if code not in zips:
                    zips.append(code)
        return zips

    def get_col(self, col_name):
        sheet = self.workbook.active
        z_index = -1
        for index,name in enumerate(sheet[1]):
            if name.value.replace(' ', '') == col_name.replace(' ', ''):
                z_index = index+1
                break
        return z_index

    def put_prices(self, name, price_list):
        sheet = self.workbook.active
        index = self.get_col(name)
        for i, price in enumerate(price_list):
            cell_str = chr(ord('A')+((index-1)%26))
            if index > 26:
                cell_str = 'A'
                cell_str += chr(ord('A')+((index-27)%26))
                
            cell_str += str(i+2)
            if price != -1:
                sheet[cell_str] = price
        self.workbook.save('output.xlsx')

    def get_addresses(self):
        zips = []
        sheet = self.workbook.active
        z_index = self.get_col("Address")
        for zipcode in sheet.iter_rows(min_row=2, min_col = z_index, max_col=z_index):
            zips.append(zipcode[0].value)
        return zips 


    def set_active_sheet(self, name):
        self.workbook.active = self.workbook[name]

    def read_zips(self):
        zips = []
        sheet = self.workbook.active
        z_index = self.get_col("Zip")
        for zipcode in sheet.iter_rows(min_row=2, min_col = z_index, max_col=z_index):
            zips.append(int(zipcode[0].value))
        return zips